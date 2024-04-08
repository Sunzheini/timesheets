from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from db.database import column_of_month_in_common_timesheets_file, column_of_year_in_common_timesheets_file, months
from support.support_functions import get_holidays_for_a_specific_month_and_year


def _determine_start_row_by_given_string(worksheet, month, year):
    """
    This function determines the start row by given month and year
    :param worksheet: the worksheet
    :param month: the month
    :param year: the year
    :return: the start row
    """
    int_year = int(year)
    capitalized_month = month.capitalize()
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                   min_col=column_of_month_in_common_timesheets_file,
                                   max_col=column_of_month_in_common_timesheets_file):
        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                continue

            if isinstance(cell_value, int):
                continue

            cell_value = cell_value.strip()
            if cell_value != capitalized_month:
                continue

            # check value of the cell in the same row but in column of year
            year_cell = worksheet[cell.row][column_of_year_in_common_timesheets_file].value
            if year_cell != int_year:
                continue
            return cell.coordinate


def get_the_info_related_to_the_employee(worksheet, coordinates):
    """
    This function gets the information related to the employee
    :param worksheet: the worksheet
    :param coordinates: the coordinates
    :return: the information related to the employee, which looks like this:
    '{'ScienceDiver': {
        1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0,
        16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0, 24: 0, 25: 0, 26: 0, 27: 0, 28: 0, 29: 0,
        30: 0, 31: 0, 'Ʃ': '=SUM(C405:AG405)'},
      '4BIZ': {
        1: 2, 2: 4, 3: 0, 4: 0, 5: 2, 6: 6, 7: 2, 8: 4, 9: 3, 10: 0, 11: 0, 12: 4, 13: 2, 14: 4, 15: 2,
        16: 3, 17: 0, 18: 0, 19: 2, 20: 4, 21: 4, 22: 0, 23: 0, 24: 0, 25: 0, 26: 4, 27: 4, 28: 8, 29: 4,
        30: 5, 31: 0, 'Ʃ': '=SUM(C406:AG406)'},
      'Bridge': {
        1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 1, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0,
        16: 1, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0, 24: 0, 25: 0, 26: 0, 27: 0, 28: 0, 29: 0,
        30: 1, 31: 0, 'Ʃ': '=SUM(C407:AG407)'},
      'InnoForward': {
        1: 6, 2: 4, 3: 0, 4: 0, 5: 6, 6: 2, 7: 6, 8: 4, 9: 4, 10: 0, 11: 0, 12: 4, 13: 6, 14: 4, 15: 6,
        16: 4, 17: 0, 18: 0, 19: 6, 20: 4, 21: 4, 22: 8, 23: 8, 24: 0, 25: 0, 26: 4, 27: 4, 28: 0, 29: 4,
        30: 2, 31: 0, 'Ʃ': '=SUM(C408:AG408)'},
      'Total Ʃ Hours ': {
        1: '=SUM(C405:C408)', 2: '=SUM(D405:D408)', 3: '=SUM(E405:E408)', 4: '=SUM(F405:F408)', 5: '=SUM(G405:G408)',
        6: '=SUM(H405:H408)', 7: '=SUM(I405:I408)', 8: '=SUM(J405:J408)', 9: '=SUM(K405:K408)', 10: '=SUM(L405:L408)',
        11: '=SUM(M405:M408)', 12: '=SUM(N405:N408)', 13: '=SUM(O405:O408)', 14: '=SUM(P405:P408)',
        15: '=SUM(Q405:Q408)', 16: '=SUM(R405:R408)', 17: '=SUM(S405:S408)', 18: '=SUM(T405:T408)',
        19: '=SUM(U405:U408)', 20: '=SUM(V405:V408)', 21: '=SUM(W405:W408)', 22: '=SUM(X405:X408)',
        23: '=SUM(Y405:Y408)', 24: '=SUM(Z405:Z408)', 25: '=SUM(AA405:AA408)', 26: '=SUM(AB405:AB408)',
        27: '=SUM(AC405:AC408)', 28: '=SUM(AD405:AD408)', 29: '=SUM(AE405:AE408)', 30: '=SUM(AF405:AF408)',
        31: '=SUM(AG405:AG408)', 'Ʃ': '=SUM(AH405:AH408)'
       }
    }'
    """
    result = {}
    to_exit = False

    # 1. Find the reference cell and the days cell ------------------------------------------------
    for row in worksheet.iter_rows(min_row=worksheet[coordinates].row + 1, max_row=worksheet.max_row,
                                   min_col=1, max_col=1):

        if to_exit:
            break

        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                continue

            if type(cell_value) == int:
                continue

            if 'Reference' not in cell_value:
                continue

            else:
                reference_cell_coordinate = cell.coordinate  # A404
                reference_cell_row = cell.row  # 404
                reference_cell_column = cell.column  # 1

                days_cell_coordinate = (
                    worksheet.cell(row=reference_cell_row - 1, column=reference_cell_column + 1).coordinate)  # B403
                days_cell_row = reference_cell_row - 1  # 403
                days_cell_column = reference_cell_column + 1  # 2

                # 1.1. Fill the result dictionary -----------------------------------------------------
                break_on_next_iter = False
                counter = 1
                while 1:
                    if break_on_next_iter:
                        break

                    current_a_column_cell = worksheet.cell(
                        row=reference_cell_row + counter, column=reference_cell_column)

                    # check if contents of reference cell include 'Total' -------------------------
                    current_project_name_column_cell_contents = current_a_column_cell.value

                    # If there is an empty project name cell after `Reference` and before `Total`, skip it
                    if current_project_name_column_cell_contents is None:
                        counter += 1
                        continue

                    if 'Total' in current_project_name_column_cell_contents:
                        break_on_next_iter = True

                    result[current_a_column_cell.value] = {}  # current_a_column_cell.value == 'ScienceDiver'

                    # for each row before empty cell in the same column ---------------------------
                    inner_counter = 1
                    while 1:

                        current_day_cell = worksheet.cell(
                            row=days_cell_row, column=days_cell_column + inner_counter)

                        # check if current_day_cell is empty, and if so, break --------------------
                        if current_day_cell.value is None:
                            break

                        current_project_cell = worksheet.cell(
                            row=reference_cell_row + counter, column=reference_cell_column + 1 + inner_counter)

                        current_project_cell_value = current_project_cell.value
                        if current_project_cell_value is None:
                            current_project_cell_value = 0

                        result[current_a_column_cell.value][current_day_cell.value] = current_project_cell_value
                        inner_counter += 1

                    counter += 1

                to_exit = True
                break

    return result


def recalculate_because_of_formulas(time_dict):
    """
    Recalculate the time dictionary because of formulas
    :param time_dict: the time dictionary
    :return: the recalculated time dictionary, which looks like this:
    '{
      'ScienceDiver': {
        1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0,
        14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0, 24: 0, 25: 0,
        26: 0, 27: 0, 28: 0, 29: 0, 30: 0, 31: 0, 'Ʃ': 0},
      '4BIZ': {
        1: 2, 2: 4, 3: 0, 4: 0, 5: 2, 6: 6, 7: 2, 8: 4, 9: 3, 10: 0, 11: 0, 12: 4, 13: 2,
        14: 4, 15: 2, 16: 3, 17: 0, 18: 0, 19: 2, 20: 4, 21: 4, 22: 0, 23: 0, 24: 0, 25: 0,
        26: 4, 27: 4, 28: 8, 29: 4, 30: 5, 31: 0, 'Ʃ': 73},
      'Bridge': {
        1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 1, 10: 0, 11: 0, 12: 0, 13: 0,
        14: 0, 15: 0, 16: 1, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0, 24: 0, 25: 0,
        26: 0, 27: 0, 28: 0, 29: 0, 30: 1, 31: 0, 'Ʃ': 3},
      'InnoForward': {
        1: 6, 2: 4, 3: 0, 4: 0, 5: 6, 6: 2, 7: 6, 8: 4, 9: 4, 10: 0, 11: 0, 12: 4, 13: 6, 14: 4,
        15: 6, 16: 4, 17: 0, 18: 0, 19: 6, 20: 4, 21: 4, 22: 8, 23: 8, 24: 0, 25: 0, 26: 4, 27: 4,
        28: 0, 29: 4, 30: 2, 31: 0, 'Ʃ': 100},
      'Total Ʃ Hours ': {
        1: 8, 2: 8, 3: 0, 4: 0, 5: 8, 6: 8, 7: 8, 8: 8, 9: 8, 10: 0, 11: 0, 12: 8, 13: 8,
        14: 8, 15: 8, 16: 8, 17: 0, 18: 0, 19: 8, 20: 8, 21: 8, 22: 8, 23: 8, 24: 0, 25: 0,
        26: 8, 27: 8, 28: 8, 29: 8, 30: 8, 31: 0, 'Ʃ': 176
      }
    }'
    """
    break_on_next_iter = False

    for project_key, value_dict in time_dict.items():
        if break_on_next_iter:
            break

        if 'Total' in project_key:
            sum_key = list(value_dict.keys())[-1]
            for day, hours in value_dict.items():
                if day == sum_key:
                    break_on_next_iter = True
                    continue

                value_dict[day] = 0
                for key, value in time_dict.items():
                    if 'Total' in key:
                        continue

                    value_dict[day] += int(value[day])

        # last item in value_dict is the sum key
        sum_key = list(value_dict.keys())[-1]
        total_sum = 0
        for day, hours in value_dict.items():
            if day == sum_key:
                continue

            total_sum += int(hours)

        time_dict[project_key][sum_key] = total_sum

    return time_dict


def read_from_excel_file_of_type_common(file_path, name, year, month):
    """
    Read from the Excel file of type common timesheets
    :param file_path: the file path of the Excel file
    :param name: the name of the employee
    :param year: the year, which as selected by the user
    :param month: the month, which as selected by the user
    :return: the information related to the employee for the given month and year in the Excel file
    """
    workbook = load_workbook(file_path)
    worksheet = workbook[name]
    coordinates = _determine_start_row_by_given_string(worksheet, month, year)  # X481
    result = get_the_info_related_to_the_employee(worksheet, coordinates)
    result = recalculate_because_of_formulas(result)
    return result


def read_from_excel_file2(file_path, sheet_name):
    """
    :param file_path: the path to the file
    :param sheet_name: the name of the sheet
    :return: the result dictionary, which looks like this:
    {
        1: 2, 2: 4, 3: 0, 4: 0, 5: 2, 6: 6, 7: 2, 8: 4, 9: 3, 10: 0, 11: 0, 12: 4, 13: 2,
        14: 4, 15: 2, 16: 3, 17: 0, 18: 0, 19: 2, 20: 4, 21: 4, 22: 0, 23: 0, 24: 0, 25: 0,
        26: 4, 27: 4, 28: 8, 29: 4, 30: 5, 31: 0, 'Ʃ': 73
    }
    """
    if len(sheet_name) == 6:
        sheet_name = '0' + sheet_name

    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    result = {}
    days_cell_row = None

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1):

        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                continue

            if isinstance(cell_value, int):
                continue

            if 'Total' not in cell_value and 'Reference' not in cell_value:
                continue

            else:
                if 'Reference' in cell_value:
                    reference_cell_coordinate = cell.coordinate
                    reference_cell_row = cell.row
                    reference_cell_column = cell.column

                    days_cell_coordinate = (
                        worksheet.cell(row=reference_cell_row - 1, column=reference_cell_column + 1).coordinate)
                    days_cell_row = reference_cell_row - 1
                    days_cell_column = reference_cell_column + 1

                else:
                    total_cell_coordinate = cell.coordinate  # A404
                    total_cell_row = cell.row  # 404
                    total_cell_column = cell.column  # 1

                    # Fill the result dictionary ---------------------------------------------------
                    counter = 1
                    while 1:
                        current_total_column_cell = worksheet.cell(
                            row=total_cell_row, column=total_cell_column + 1 + counter)     # C18

                        # check empty -------------------------------------------------------------
                        if current_total_column_cell.value is None:
                            break

                        if counter == 32:
                            sum_of_all_values = 0
                            for key, value in result.items():
                                sum_of_all_values += value

                            result['Ʃ'] = sum_of_all_values
                            break

                        # sum values from the same column from days_cell_row to total_cell_row-1 --
                        start_cell = worksheet.cell(row=days_cell_row + 2, column=total_cell_column + 1 + counter)
                        end_cell = worksheet.cell(row=total_cell_row - 1, column=total_cell_column + 1 + counter)
                        all_cells = worksheet[start_cell.coordinate:end_cell.coordinate]

                        sum_of_values = 0
                        for current_cell in all_cells:
                            the_cell = current_cell[0]
                            if the_cell.value is None:
                                continue

                            sum_of_values += the_cell.value

                        result[counter] = sum_of_values
                        counter += 1

    return result


def read_from_excel_file_and_insert_rows(file_path, sheet_name, year, month, data, dict_with_holidays):
    """
    :param dict_with_holidays: looks like this:
    {'1': 'work day', '2': 'work day', '3': 'work day', '4': 'work day', '5': 'weekend',}
    :param data: the data for the given project to insert, looks like this:
    {
        1: 6, 2: 4, 3: 0, 4: 0, 5: 6, 6: 2, 7: 6, 8: 4, 9: 4, 10: 0, 11: 0, 12: 4, 13: 6, 14: 4,
        15: 6, 16: 4, 17: 0, 18: 0, 19: 6, 20: 4, 21: 4, 22: 8, 23: 8, 24: 0, 25: 0, 26: 4, 27: 4,
        28: 0, 29: 4, 30: 2, 31: 0, 'Ʃ': 100
    }
    :param month: the month
    :param year: the year
    :param file_path: the path to the file
    :param sheet_name: the name of the sheet
    :return: a string indicating if the operation was successful or not
    """
    result = ''

    if len(sheet_name) == 6:
        sheet_name = '0' + sheet_name

    workbook = load_workbook(file_path)

    # 1. check if a worksheet exists with the same name --------------------------------------------
    if sheet_name in workbook.sheetnames:
        result = f'The sheet "{sheet_name}" already exists in the file "{file_path}"'
        return result

    # 2. Copy from a template sheet and rename it --------------------------------------------------
    template_sheet = workbook['template']
    new_sheet = workbook.copy_worksheet(template_sheet)
    new_sheet.title = sheet_name

    # 3. Update the year and month in the new sheet ------------------------------------------------
    new_sheet['X1'] = month.capitalize()
    new_sheet['AE1'] = year

    # 4. Find the reference cell and the days cell -------------------------------------------------
    days_cell_row = None

    for row in new_sheet.iter_rows(min_row=1, max_row=new_sheet.max_row, min_col=1, max_col=1):

        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                continue

            if isinstance(cell_value, int):
                continue

            if 'Total' not in cell_value and 'Reference' not in cell_value:
                continue

            else:
                if 'Reference' in cell_value:
                    reference_cell_coordinate = cell.coordinate
                    reference_cell_row = cell.row
                    reference_cell_column = cell.column

                    days_cell_coordinate = (
                        new_sheet.cell(row=reference_cell_row - 1, column=reference_cell_column + 1).coordinate)
                    days_cell_row = reference_cell_row - 1
                    days_cell_column = reference_cell_column + 1

                else:   # 'Total' in cell_value
                    total_cell_coordinate = cell.coordinate  # A404
                    total_cell_row = cell.row  # 404
                    total_cell_column = cell.column  # 1

    # copy the data from the data dictionary to the row before the total row -------------------------
                    for key, value in data.items():
                        if key == 'Ʃ':
                            break

                        new_sheet.cell(row=total_cell_row + 2, column=key+2, value=value)

                    # color the weekend cells ---------------------------------------------------------------------
                    for key, value in dict_with_holidays.items():
                        if value == 'weekend':
                            for current_row in new_sheet.iter_rows(min_row=days_cell_row + 2,
                                                                   max_row=total_cell_row - 1,
                                                                   min_col=int(key) + 2, max_col=int(key) + 2):
                                for current_cell in current_row:
                                    current_cell.fill = PatternFill(start_color='D9D9D9',
                                                                    end_color='D9D9D9',
                                                                    fill_type='solid')

    result = f'The sheet "{sheet_name}" was successfully created in the file "{file_path}"'

    # Save the workbook after copying and renaming the sheet
    workbook.save(file_path)

    return result
